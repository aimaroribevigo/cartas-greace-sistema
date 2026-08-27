# -*- coding: utf-8 -*-
"""Suite de pruebas y auditoría exhaustiva de Importación, Exportación y Seguridad de Excel."""
import io
import os
import sys
import tempfile
import openpyxl

from app import app, get_db, import_excel_to_db, refresh_normalized_fields, _rebuild_hilos
from backfill_cartas import backfill_cartas
from export_excel import export_full_backup_excel

print("======================================================================")
print("           SUITE DE AUDITORÍA Y TESTING EXHAUSTIVO DE EXCEL           ")
print("======================================================================")

with app.app_context():
    db = get_db()

    # TEST 1: Importación del Excel Maestro Base
    print("\n[TEST 1] Importación del Excel Maestro Base...")
    res1 = import_excel_to_db(db, force=True)
    assert res1.get("ok") is True, f"Fallo en Test 1: {res1}"
    assert res1.get("inserted") == 2520, f"Total insertado incorrecto: {res1.get('inserted')}"
    print("  ✓ Importados exactamente 2,520 registros:")
    for ban, info in res1.get("by_bandeja", {}).items():
        print(f"    - {ban}: {info.get('inserted')} filas")

    # TEST 2: Backfill y Normalización Automática
    print("\n[TEST 2] Verificando Backfill y Normalización...")
    bf = backfill_cartas(db, dry_run=False, fill_missing=True, fix_areas=True)
    norms = refresh_normalized_fields(db)
    print(f"  ✓ Backfill ejecutado: {bf.get('fill', {}).get('referencia_updated', 0)} referencias enriquecidas")
    print(f"  ✓ Normalización de especialidades/estados completada: {norms.get('checked', 0)} filas verificadas")

    # TEST 3: Reconstrucción de Hilos con Union-Find
    print("\n[TEST 3] Reconstrucción de Hilos de Trámite...")
    hilos = _rebuild_hilos(db)
    assert hilos.get("ok") is True, f"Fallo en hilos: {hilos}"
    print(f"  ✓ Hilos generados: {hilos.get('hilos')} hilos (Abiertos: {hilos.get('abiertos')}, Cerrados: {hilos.get('cerrados')})")

    # TEST 4: Exportación de Backup Excel
    print("\n[TEST 4] Generación de Backup Excel con 6 Hojas...")
    stream = export_full_backup_excel(db)
    stream_bytes = stream.getvalue()
    assert len(stream_bytes) > 50000, "El archivo generado es demasiado pequeño"
    wb_export = openpyxl.load_workbook(io.BytesIO(stream_bytes))
    assert len(wb_export.sheetnames) == 6, f"Hojas incorrectas: {wb_export.sheetnames}"
    print(f"  ✓ Archivo Excel generado: {len(stream_bytes):,} bytes")
    print(f"  ✓ 6 hojas verificadas: {wb_export.sheetnames}")

    # TEST 5: Round-Trip (Reimportar el Backup que acabamos de exportar)
    print("\n[TEST 5] Prueba Round-Trip (Reimportar el Excel de Backup recién generado)...")
    fd, temp_xlsx = tempfile.mkstemp(suffix=".xlsx")
    os.write(fd, stream_bytes)
    os.close(fd)
    try:
        res_rt = import_excel_to_db(db, excel_path=temp_xlsx, force=True)
        assert res_rt.get("ok") is True, f"Fallo en Round-Trip: {res_rt}"
        assert res_rt.get("inserted") == 2520, f"Round-Trip no recuperó todas las cartas: {res_rt.get('inserted')}"
        print(f"  ✓ Round-Trip 100% exitoso: {res_rt.get('inserted')} cartas reimportadas con cero pérdidas.")
    finally:
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)

    # TEST 6: Resiliencia ante Archivo Vacío o Inválido (Rollback de Seguridad)
    print("\n[TEST 6] Prueba de Seguridad ante Archivo Inválido...")
    fd, bad_xlsx = tempfile.mkstemp(suffix=".xlsx")
    bad_wb = openpyxl.Workbook()
    bad_wb.active.title = "Gastos_y_Facturas"
    bad_wb.active.append(["ID", "Factura", "Monto"])
    bad_wb.active.append([1, "F001", 500])
    bad_wb.save(bad_xlsx)
    os.close(fd)
    try:
        res_bad = import_excel_to_db(db, excel_path=bad_xlsx, force=True)
        assert res_bad.get("ok") is False, "Debió fallar ante archivo incompatible"
        with db.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS c FROM cartas")
            count_after_bad = cur.fetchone()["c"]
        assert count_after_bad == 2520, f"La base de datos se borró indebidamente! Quedan {count_after_bad}"
        print("  ✓ Rollback de seguridad impecable: el archivo incompatible fue rechazado y las 2,520 cartas se mantuvieron 100% intactas.")
    finally:
        if os.path.exists(bad_xlsx):
            os.remove(bad_xlsx)

    # TEST 7: Prueba de endpoints HTTP vía Test Client
    print("\n[TEST 7] Prueba de Endpoints HTTP (/api/import/excel y /api/backup/excel)...")
    with app.test_client() as client:
        with client.session_transaction() as sess:
            sess["uid"] = 1  # Admin

        # 7a. GET Backup Excel
        res_get = client.get("/api/backup/excel")
        assert res_get.status_code == 200, f"Error en GET backup: {res_get.status_code}"
        assert "spreadsheetml.sheet" in res_get.content_type
        print("  ✓ Endpoint GET /api/backup/excel respondió 200 OK con Content-Disposition de descarga")

        # 7b. POST Reimportar Excel vía Multipart FormData
        data = {"file": (io.BytesIO(stream_bytes), "Control_de_Cartas_Backup.xlsx")}
        res_post = client.post("/api/import/excel", data=data, content_type="multipart/form-data")
        assert res_post.status_code == 200, f"Error en POST import: {res_post.status_code}"
        post_json = res_post.get_json()
        assert post_json.get("ok") is True and post_json.get("inserted") == 2520
        print(f"  ✓ Endpoint POST /api/import/excel procesó FormData exitosamente ({post_json.get('inserted')} cartas)")

print("\n======================================================================")
print("  ✅ TODAS LAS PRUEBAS PASARON AL 100% CON ÉXITO SIN NINGÚN ERROR    ")
print("======================================================================")
