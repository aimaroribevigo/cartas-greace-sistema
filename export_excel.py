# -*- coding: utf-8 -*-
"""Generador de Backup Excel (.xlsx) completo y compatible con el importador."""
from __future__ import annotations

import io
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET_CONFIGS = [
    {
        "bandeja": "residente",
        "sheet_name": "1.Cartas. Res.",
        "title": "REGISTRO DE CARTAS EMITIDAS POR RESIDENCIA DE OBRA",
        "headers": [
            ("N°", "n_orden", 8, "center"),
            ("FECHA DE EMISIÓN", "fecha", 16, "center"),
            ("N° DE CARTA", "n_documento", 28, "left"),
            ("ASUNTO", "asunto", 55, "left"),
            ("ESP. RESP.", "especialidad", 22, "center"),
            ("ESTADO DE RESPUESTA", "estado", 22, "center"),
            ("REFERENCIAS", "referencias", 30, "left"),
            ("FOLIOS", "folios", 10, "center"),
            ("CD", "cd", 8, "center"),
            ("DIRIGIDO A:", "dirigido_a", 28, "left"),
            ("RECEPTOR A:", "receptor", 28, "left"),
            ("CARGO", "cargo", 25, "left"),
            ("OBSERVACIÓN", "observacion", 45, "left"),
        ],
    },
    {
        "bandeja": "rl",
        "sheet_name": "2.Cart. RL",
        "title": "REGISTRO DE CARTAS EMITIDAS POR REPRESENTANTE LEGAL",
        "headers": [
            ("N°", "n_orden", 8, "center"),
            ("FECHA DE EMISIÓN", "fecha", 16, "center"),
            ("N° DE DOCUMENTO", "n_documento", 28, "left"),
            ("ASUNTO", "asunto", 55, "left"),
            ("ESP. RESP.", "especialidad", 22, "center"),
            ("ESTADO DE RESPUESTA", "estado", 22, "center"),
            ("REFERENCIAS", "referencias", 30, "left"),
            ("FOLIOS", "folios", 10, "center"),
            ("CD", "cd", 8, "center"),
            ("DIRIGIDO A:", "dirigido_a", 28, "left"),
            ("RECEPTOR A:", "receptor", 28, "left"),
            ("CARGO", "cargo", 25, "left"),
            ("OBSERVACION", "observacion", 45, "left"),
        ],
    },
    {
        "bandeja": "recibida_sup",
        "sheet_name": "3.Cart.Recb.Sup.",
        "title": "REGISTRO DE CARTAS RECIBIDAS DE LA SUPERVISIÓN",
        "headers": [
            ("N°", "n_orden", 8, "center"),
            ("FECHA DE RECEPCIÓN", "fecha", 16, "center"),
            ("N° DE DOCUMENTO", "n_documento", 28, "left"),
            ("ASUNTO", "asunto", 55, "left"),
            ("EMPRESA", "empresa", 22, "left"),
            ("AREA", "area", 24, "center"),
            ("ESP. RESP.", "especialidad", 22, "center"),
            ("ESTADO DE RESPUESTA", "estado", 22, "center"),
            ("REFERENCIAS", "referencias", 30, "left"),
            ("FOLIOS", "folios", 10, "center"),
            ("CD", "cd", 8, "center"),
            ("DIRIGIDO A:", "dirigido_a", 28, "left"),
            ("RECEPTOR A:", "receptor", 28, "left"),
            ("CARGO", "cargo", 25, "left"),
            ("FECHA RESPUESTA", "fecha_respuesta", 16, "center"),
            ("CARTA DE RESPUESTA", "carta_respuesta", 28, "left"),
            ("OBSERVACIÓN", "observacion", 45, "left"),
        ],
    },
    {
        "bandeja": "recibida_otros",
        "sheet_name": "4.Cart.Recb.Otros",
        "title": "REGISTRO DE CARTAS RECIBIDAS DE OTROS / JRD",
        "headers": [
            ("N°", "n_orden", 8, "center"),
            ("FECHA DE RECEPCIÓN", "fecha", 16, "center"),
            ("N° DE DOCUMENTO", "n_documento", 28, "left"),
            ("ASUNTO", "asunto", 55, "left"),
            ("ESP. RESP.", "especialidad", 22, "center"),
            ("ESTADO DE RESPUESTA", "estado", 22, "center"),
            ("REFERENCIAS", "referencias", 30, "left"),
            ("FOLIOS", "folios", 10, "center"),
            ("CD", "cd", 8, "center"),
            ("DIRIGIDO A:", "dirigido_a", 28, "left"),
            ("RECEPTOR A:", "receptor", 28, "left"),
            ("CARGO", "cargo", 25, "left"),
            ("OBSERVACION", "observacion", 45, "left"),
        ],
    },
    {
        "bandeja": "recibida_pronis",
        "sheet_name": "5.Cartas Recibidas Pronis ",
        "title": "REGISTRO DE CARTAS RECIBIDAS DE PRONIS / MINSA",
        "headers": [
            ("N°", "n_orden", 8, "center"),
            ("FECHA DE RECEPCION", "fecha", 16, "center"),
            ("N° DE DOCUMENTO", "n_documento", 28, "left"),
            ("ASUNTO", "asunto", 55, "left"),
            ("ESP. RESP.", "especialidad", 22, "center"),
            ("ESTADO DE RESPUESTA", "estado", 22, "center"),
            ("REFERENCIAS", "referencias", 30, "left"),
            ("FOLIOS", "folios", 10, "center"),
            ("DIRIGIDO A:", "dirigido_a", 28, "left"),
            ("RECEPTOR A:", "receptor", 28, "left"),
            ("OBSERVACIONES", "observacion", 45, "left"),
        ],
    },
    {
        "bandeja": "recibida_mpsc",
        "sheet_name": "6.Cartas Recibidas MPSC  ",
        "title": "REGISTRO DE CARTAS RECIBIDAS DE LA MUNICIPALIDAD (MPSC)",
        "headers": [
            ("N°", "n_orden", 8, "center"),
            ("FECHA DE RECEPCION", "fecha", 16, "center"),
            ("N° DE DOCUMENTO", "n_documento", 28, "left"),
            ("ASUNTO", "asunto", 55, "left"),
            ("ESP. RESP.", "especialidad", 22, "center"),
            ("ESTADO DE RESPUESTA", "estado", 22, "center"),
            ("REFERENCIAS", "referencias", 30, "left"),
            ("FOLIOS", "folios", 10, "center"),
            ("DIRIGIDO A:", "dirigido_a", 28, "left"),
            ("RECEPTOR A:", "receptor", 28, "left"),
            ("OBSERVACIONES", "observacion", 45, "left"),
        ],
    },
]

# Estilos profesionales
COLOR_HEADER_BG = "1F4E79"       # Azul corporativo oscuro
COLOR_HEADER_TXT = "FFFFFF"      # Blanco
COLOR_ZEBRA = "F2F5F9"           # Alternado suave
COLOR_BORDER = "D9D9D9"          # Gris claro

font_title = Font(name="Calibri", size=14, bold=True, color="1F4E79")
font_meta = Font(name="Calibri", size=10, italic=True, color="595959")
font_header = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_TXT)
font_data = Font(name="Calibri", size=10, color="000000")

fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)


def export_full_backup_excel(conn) -> io.BytesIO:
    """Genera un archivo Excel (.xlsx) con todas las cartas en 6 hojas estructuradas."""
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
                id, bandeja, sentido, n_orden, fecha, n_documento, tipo_documento,
                asunto, especialidad, especialidad_norm, estado, estado_norm,
                referencias, referencia, folios, cd, dirigido_a, receptor, cargo,
                observacion, area, empresa, caducidad, fecha_respuesta, carta_respuesta,
                creado_en, actualizado_en, hilo_id
            FROM cartas
            ORDER BY bandeja, COALESCE(n_orden, 999999), fecha, id
            """
        )
        all_cartas = cur.fetchall()

    cartas_by_bandeja: dict[str, list[dict]] = {}
    for c in all_cartas:
        b = str(c.get("bandeja") or "residente").strip().lower()
        cartas_by_bandeja.setdefault(b, []).append(c)

    wb = openpyxl.Workbook()
    # Eliminar hoja default
    wb.remove(wb.active)

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    for cfg in SHEET_CONFIGS:
        ws = wb.create_sheet(title=cfg["sheet_name"])
        ws.views.sheetView[0].showGridLines = True

        # Fila 1: Título
        ws.cell(row=1, column=1, value=cfg["title"]).font = font_title
        # Fila 2: Subtítulo con fecha
        ws.cell(row=2, column=1, value=f"Backup generado el: {now_str} | Sistema Greace HLP").font = font_meta

        # Fila 4: Encabezados
        header_row = 4
        ws.row_dimensions[header_row].height = 26

        for col_idx, (hdr_name, _, width, align) in enumerate(cfg["headers"], start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=hdr_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(width, len(hdr_name) + 3)

        # Filas de Datos
        bandeja_key = cfg["bandeja"]
        rows = cartas_by_bandeja.get(bandeja_key, [])
        data_start_row = 5

        for r_idx, carta in enumerate(rows, start=data_start_row):
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)

            for col_idx, (_, field_key, _, align) in enumerate(cfg["headers"], start=1):
                val = carta.get(field_key)
                # Formatear fechas
                if isinstance(val, (datetime, date)):
                    val = val.strftime("%Y-%m-%d")

                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.font = font_data
                cell.border = thin_border

                if is_even:
                    cell.fill = fill_zebra

                if align == "center":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif align == "right":
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    # Texto normal con ajuste en asunto/observaciones
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(field_key in ("asunto", "observacion")))

        # Habilitar autofiltro en la fila de encabezados
        if rows:
            max_col_letter = get_column_letter(len(cfg["headers"]))
            ws.auto_filter.ref = f"A{header_row}:{max_col_letter}{data_start_row + len(rows) - 1}"

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
