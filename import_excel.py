# -*- coding: utf-8 -*-
"""Importación del Excel Control de Cartas → tabla MySQL `cartas`."""
from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path

import openpyxl

from normalizers import normalize_especialidad, normalize_estado

DEFAULT_EXCEL = Path(__file__).resolve().parent / "Control_de_Cartas_2025_HLP_Mejorado.xlsx"
BATCH_SIZE = 500
EMPTY_STREAK_STOP = 80

SHEETS = [
    {
        "bandeja": "residente",
        "sheet": "1.Cartas. Res.",
        "header_row": 19,
        "layout": "emitida",
        "sentido": "emitida",
        "max_col": 14,
    },
    {
        "bandeja": "rl",
        "sheet": "2.Cart. RL",
        "header_row": 19,
        "layout": "emitida",
        "sentido": "emitida",
        "max_col": 14,
    },
    {
        "bandeja": "recibida_sup",
        "sheet": "3.Cart.Recb.Sup.",
        "header_row": 14,
        "layout": "recibida_sup",
        "sentido": "recibida",
        "max_col": 19,
    },
    {
        "bandeja": "recibida_otros",
        "sheet": "4.Cart.Recb.Otros",
        "header_row": 14,
        "layout": "recibida_otros",
        "sentido": "recibida",
        "max_col": 14,
    },
    {
        "bandeja": "recibida_pronis",
        "sheet": "5.Cartas Recibidas Pronis ",
        "header_row": 16,
        "layout": "recibida_simple",
        "sentido": "recibida",
        "max_col": 12,
    },
    {
        "bandeja": "recibida_mpsc",
        "sheet": "6.Cartas Recibidas MPSC  ",
        "header_row": 13,
        "layout": "recibida_mpsc",
        "sentido": "recibida",
        "max_col": 12,
    },
]

INSERT_SQL = """
    INSERT INTO cartas (
        bandeja, sentido, n_orden, fecha, n_documento, asunto,
        especialidad, especialidad_norm, estado, estado_norm,
        referencias, folios, cd, dirigido_a, receptor, cargo,
        observacion, area, empresa, caducidad, fecha_respuesta, carta_respuesta
    ) VALUES (
        %s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s
    )
"""


def _as_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = s[:10]
    for fmt in ("%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def _as_str(v, max_len=None):
    if v is None:
        return None
    s = str(v).replace("\r\n", "\n").strip()
    if not s:
        return None
    if max_len:
        return s[:max_len]
    return s


def _as_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _val(cells, idx):
    """cells es 0-based; idx es 1-based (columna Excel)."""
    i = idx - 1
    if i < 0 or i >= len(cells):
        return None
    return cells[i]


def parse_cells(layout: str, cells) -> dict | None:
    n_documento = _as_str(_val(cells, 3), 255)
    if not n_documento:
        return None
    up = n_documento.upper()
    if up in ("N° DE DOCUMENTO", "Nº DE DOCUMENTO", "TOTAL", "N°"):
        return None

    base = {
        "n_orden": _as_int(_val(cells, 1)),
        "fecha": _as_date(_val(cells, 2)),
        "n_documento": n_documento,
        "asunto": _as_str(_val(cells, 5)),
        "empresa": None,
        "area": None,
        "especialidad": None,
        "estado": None,
        "referencias": None,
        "folios": None,
        "cd": None,
        "dirigido_a": None,
        "receptor": None,
        "cargo": None,
        "observacion": None,
        "caducidad": None,
        "fecha_respuesta": None,
        "carta_respuesta": None,
    }

    if layout == "emitida":
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "cd": _as_str(_val(cells, 10), 20),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
                "cargo": _as_str(_val(cells, 13), 255),
                "observacion": _as_str(_val(cells, 14)),
            }
        )
    elif layout == "recibida_sup":
        base.update(
            {
                "area": _as_str(_val(cells, 9), 255),
                "especialidad": _as_str(_val(cells, 10), 255),
                "estado": _as_str(_val(cells, 11), 120),
                "referencias": _as_str(_val(cells, 12)),
                "folios": _as_str(_val(cells, 13), 50),
                "cd": _as_str(_val(cells, 14), 20),
                "dirigido_a": _as_str(_val(cells, 15), 255),
                "receptor": _as_str(_val(cells, 16), 255),
                "cargo": _as_str(_val(cells, 17), 255),
                "fecha_respuesta": _as_date(_val(cells, 18)),
                "carta_respuesta": _as_str(_val(cells, 19), 255),
            }
        )
    elif layout == "recibida_simple":
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "observacion": _as_str(_val(cells, 10)),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
            }
        )
    elif layout == "recibida_otros":
        # Hoja 4: A N°, B fecha, C doc, D vacío, E asunto, F esp, G estado…
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "cd": _as_str(_val(cells, 10), 20),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
                "cargo": _as_str(_val(cells, 13), 255),
                "observacion": _as_str(_val(cells, 14)),
            }
        )
    elif layout == "recibida_mpsc":
        # Misma grilla que simple; estado suele venir vacío → no inventar estado
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "observacion": _as_str(_val(cells, 10)),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
            }
        )
        # Asuntos de solo conocimiento
        asunto_u = (base.get("asunto") or "").upper()
        if not base.get("estado") and any(
            k in asunto_u for k in ("HACE DE CONOCIMIENTO", "HACE LLEGAR COPIA", "AUTORIZACIÓN", "AUTORIZACION")
        ):
            base["estado"] = "PARA CONOCIMIENTO"
    else:
        return None

    base["estado_norm"] = normalize_estado(base["estado"])
    base["especialidad_norm"] = normalize_especialidad(base["especialidad"])
    return base


def _row_tuple(cfg, row):
    return (
        cfg["bandeja"],
        cfg["sentido"],
        row["n_orden"],
        row["fecha"],
        row["n_documento"],
        row["asunto"],
        row["especialidad"],
        row["especialidad_norm"],
        row["estado"],
        row["estado_norm"],
        row["referencias"],
        row["folios"],
        row["cd"],
        row["dirigido_a"],
        row["receptor"],
        row["cargo"],
        row["observacion"],
        row["area"],
        row["empresa"],
        row["caducidad"],
        row["fecha_respuesta"],
        row["carta_respuesta"],
    )


def iter_parsed_rows(ws, header_row: int, layout: str, max_col: int):
    empty = 0
    for row_idx, cells in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_col=max_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        row = parse_cells(layout, cells or ())
        if not row:
            empty += 1
            if empty >= EMPTY_STREAK_STOP:
                break
            continue
        empty = 0
        yield row_idx, row


def import_excel_to_db(conn, excel_path: Path | None = None, force: bool = False) -> dict:
    excel_path = Path(
        excel_path
        or os.environ.get("EXCEL_PATH")
        or DEFAULT_EXCEL
    )
    if not excel_path.exists():
        return {"ok": False, "error": f"Excel no encontrado: {excel_path}"}

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) AS c FROM cartas")
        count = cur.fetchone()["c"]
        if count > 0 and not force:
            return {"ok": True, "skipped": True, "existing": count}

        if force and count > 0:
            cur.execute("DELETE FROM cartas")

        # read_only + values_only: evita el freeze por ws.cell / max_row
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        inserted = 0
        by_bandeja = {}
        try:
            for cfg in SHEETS:
                name = cfg["sheet"]
                if name not in wb.sheetnames:
                    match = next((s for s in wb.sheetnames if s.strip() == name.strip()), None)
                    if not match:
                        by_bandeja[cfg["bandeja"]] = {"error": "hoja no encontrada"}
                        continue
                    name = match
                ws = wb[name]
                batch = []
                n = 0
                for _, row in iter_parsed_rows(ws, cfg["header_row"], cfg["layout"], cfg["max_col"]):
                    batch.append(_row_tuple(cfg, row))
                    if len(batch) >= BATCH_SIZE:
                        cur.executemany(INSERT_SQL, batch)
                        n += len(batch)
                        inserted += len(batch)
                        batch.clear()
                if batch:
                    cur.executemany(INSERT_SQL, batch)
                    n += len(batch)
                    inserted += len(batch)
                by_bandeja[cfg["bandeja"]] = {"inserted": n}
                print(f"[import] {cfg['bandeja']}: {n} filas", flush=True)
        finally:
            wb.close()

        conn.commit()
        return {
            "ok": True,
            "inserted": inserted,
            "by_bandeja": by_bandeja,
            "excel": str(excel_path),
        }
