# -*- coding: utf-8 -*-
"""Importación inteligente del Excel Control de Cartas → tabla MySQL `cartas`."""
from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path

import openpyxl

from core.normalizers import infer_estado_from_row, normalize_especialidad, normalize_estado
from services.backfill_cartas import fix_areas_responsables

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = BASE_DIR / "Control_de_Cartas_2025_HLP_Mejorado.xlsx"
if not DEFAULT_EXCEL.exists():
    DEFAULT_EXCEL = BASE_DIR / "data" / "Control_de_Cartas_2025_HLP_Mejorado.xlsx"

BATCH_SIZE = 500
EMPTY_STREAK_STOP = 80

SHEETS = [
    {
        "bandeja": "residente",
        "sheet": "1.Cartas. Res.",
        "keywords": ["residente", "res.", "cartas. res", "cartas res", "ro", "1."],
        "header_row": 19,
        "layout": "emitida",
        "sentido": "emitida",
        "max_col": 14,
    },
    {
        "bandeja": "rl",
        "sheet": "2.Cart. RL",
        "keywords": ["cart. rl", "cartas rl", "rl", "legal", "2."],
        "header_row": 19,
        "layout": "emitida",
        "sentido": "emitida",
        "max_col": 14,
    },
    {
        "bandeja": "recibida_sup",
        "sheet": "3.Cart.Recb.Sup.",
        "keywords": ["recb.sup", "supervis", "recibida sup", "recibidas sup", "3."],
        "header_row": 14,
        "layout": "recibida_sup",
        "sentido": "recibida",
        "max_col": 19,
    },
    {
        "bandeja": "recibida_otros",
        "sheet": "4.Cart.Recb.Otros",
        "keywords": ["recb.otros", "otros", "jrd", "recibidas otros", "4."],
        "header_row": 14,
        "layout": "recibida_otros",
        "sentido": "recibida",
        "max_col": 14,
    },
    {
        "bandeja": "recibida_pronis",
        "sheet": "5.Cartas Recibidas Pronis ",
        "keywords": ["pronis", "entidad", "minsa", "recibidas pronis", "5."],
        "header_row": 16,
        "layout": "recibida_simple",
        "sentido": "recibida",
        "max_col": 12,
    },
    {
        "bandeja": "recibida_mpsc",
        "sheet": "6.Cartas Recibidas MPSC  ",
        "keywords": ["mpsc", "muni", "municipalidad", "recibidas mpsc", "6."],
        "header_row": 13,
        "layout": "recibida_mpsc",
        "sentido": "recibida",
        "max_col": 12,
    },
]

INSERT_SQL = """
    INSERT INTO cartas (
        bandeja, sentido, n_orden, fecha, n_documento, tipo_documento, asunto,
        especialidad, especialidad_norm, estado, estado_norm,
        referencias, folios, cd, dirigido_a, receptor, cargo,
        observacion, area, empresa, caducidad, fecha_respuesta, carta_respuesta
    ) VALUES (
        %s,%s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s,
        %s,%s,%s,%s,%s,%s
    )
"""


def infer_tipo_documento(doc_str: str | None, asunto: str | None = None) -> str:
    blob = f"{doc_str or ''} {asunto or ''}".upper()
    if not blob.strip():
        return "CARTA"
    if "FICHA" in blob or ("MATERIAL" in blob and any(k in blob for k in ("APROBAC", "PRESENTAC", "MUESTRA"))):
        return "FICHA TÉCNICA"
    if "CONSULTA" in blob or "RFI" in blob or "INCOMPATIBILIDAD" in blob or "ACLARAC" in blob:
        return "CONSULTA"
    if "VALORIZAC" in blob:
        return "VALORIZACIÓN"
    if "PLANO" in blob:
        return "PLANOS"
    if "INFORME" in blob:
        return "INFORME"
    if "OFICIO" in blob:
        return "OFICIO"
    if "ASIENTO" in blob or "CUADERNO" in blob:
        return "ASIENTO DE CUADERNO"
    if "MEMO" in blob:
        return "MEMORANDO"
    if "NOTARIAL" in blob:
        return "CARTA NOTARIAL"
    if "ACTA" in blob:
        return "ACTA"
    if "SOLICITUD" in blob or ("AMPLIAC" in blob and "PLAZO" in blob):
        return "SOLICITUD"
    return "CARTA"


def _as_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            if 10000 <= float(v) <= 80000:
                from datetime import timedelta
                return date(1899, 12, 30) + timedelta(days=int(float(v)))
        except Exception:
            pass
    s = str(v).strip()
    if s.isdigit():
        try:
            num = int(s)
            if 10000 <= num <= 80000:
                from datetime import timedelta
                return date(1899, 12, 30) + timedelta(days=num)
        except Exception:
            pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.split("T")[0]).date()
    except ValueError:
        return None


def _as_str(v, max_len=None):
    if v is None:
        return None
    s = str(v).replace("\r\n", "\n").strip()
    if not s:
        return None
    if max_len:
        return s[:max_len]
    return s


def _as_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def _val(cells, idx):
    """cells es 0-based; idx es 1-based (columna Excel)."""
    i = idx - 1
    if i < 0 or i >= len(cells):
        return None
    return cells[i]


def find_matching_sheet(wb_sheetnames: list[str], cfg: dict) -> str | None:
    target = cfg["sheet"].strip().lower()
    for s in wb_sheetnames:
        if s.strip().lower() == target:
            return s
    for s in wb_sheetnames:
        s_low = s.strip().lower()
        if any(k in s_low for k in cfg.get("keywords", [])):
            return s
    return None


def detect_header_row_and_map(ws, default_row: int, max_scan: int = 35) -> tuple[int, dict[str, int] | None]:
    """Detecta automáticamente la fila de encabezados y el mapeo de columnas."""
    header_keywords = {
        "n_orden": ["n°", "nro", "item", "orden", "nº", "n° orden", "n° de orden", "no."],
        "fecha": ["fecha", "fec", "f. emision", "f.emision", "fecha doc", "date", "fecha de emisión", "fecha emision", "fec. doc"],
        "n_documento": ["documento", "n° de documento", "nº de documento", "n° doc", "carta", "n_documento", "nro doc", "documento n°", "n° carta", "nº carta", "código", "codigo", "doc", "n° doc.", "documento de referencia"],
        "asunto": ["asunto", "descripcion", "referencia", "detalle", "resumen", "tema", "subject", "descripción", "contenido", "asunto / descripción"],
        "especialidad": ["especialidad", "esp", "disciplina", "specialty", "área técnica", "area tecnica"],
        "estado": ["estado", "situacion", "status", "situación", "estado del trámite", "estado tramite"],
        "referencias": ["referencias", "antecedente", "antecedentes", "ref.", "referencia", "doc. antecedente", "carta antecedente"],
        "folios": ["folio", "folios", "n° folios"],
        "cd": ["cd", "adjunto", "disco", "anexos"],
        "dirigido_a": ["dirigido", "dirigido a", "destinatario", "para", "dirigido_a", "to", "dirigido a:"],
        "receptor": ["receptor", "emisor", "de", "remitente", "from", "de:"],
        "cargo": ["cargo", "puesto", "cargo del destinatario"],
        "observacion": ["observacion", "observaciones", "obs", "comentario", "notas", "observación"],
        "area": ["area", "responsable", "asignado", "especialista", "area interna", "área", "área responsable"],
        "empresa": ["empresa", "consorcio", "entidad"],
        "fecha_respuesta": ["fecha respuesta", "fec. rpta", "fecha rpta", "f. rpta", "fecha de respuesta"],
        "carta_respuesta": ["carta respuesta", "carta rpta", "doc respuesta", "doc rpta", "documento de respuesta"],
        "bandeja": ["bandeja", "origen", "tipo bandeja", "tipo de carta"],
    }

    best_row = default_row
    best_map: dict[str, int] | None = None
    best_score = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        if not row:
            continue
        cur_map = {}
        score = 0
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            txt = str(cell).strip().lower()
            for field, kw_list in header_keywords.items():
                if field not in cur_map and any(k == txt or k in txt for k in kw_list):
                    cur_map[field] = col_idx
                    score += 1
                    break
        if score >= 2 and ("n_documento" in cur_map or "asunto" in cur_map or "fecha" in cur_map):
            if score > best_score:
                best_score = score
                best_row = r_idx
                best_map = cur_map

    if best_score >= 2 and best_map:
        return best_row, best_map
    return default_row, None


def parse_cells_with_map(cells: tuple, col_map: dict[str, int]) -> dict | None:
    def g(field, max_len=None):
        idx = col_map.get(field)
        if idx is None or idx >= len(cells):
            return None
        return _as_str(cells[idx], max_len)

    def g_date(field):
        idx = col_map.get(field)
        if idx is None or idx >= len(cells):
            return None
        return _as_date(cells[idx])

    def g_int(field):
        idx = col_map.get(field)
        if idx is None or idx >= len(cells):
            return None
        return _as_int(cells[idx])

    n_doc = g("n_documento", 255)
    asunto = g("asunto")
    if not n_doc and not asunto:
        return None

    if n_doc and n_doc.upper() in ("TOTAL", "N°", "NRO", "N° DE DOCUMENTO", "DOCUMENTO", "ITEM"):
        return None

    estado_raw = g("estado", 120)
    esp_raw = g("especialidad", 255)

    base = {
        "n_orden": g_int("n_orden"),
        "fecha": g_date("fecha"),
        "n_documento": n_doc or "S/N",
        "asunto": asunto,
        "especialidad": esp_raw,
        "especialidad_norm": normalize_especialidad(esp_raw),
        "estado": estado_raw,
        "estado_norm": normalize_estado(estado_raw),
        "referencias": g("referencias"),
        "folios": g("folios", 50),
        "cd": g("cd", 20),
        "dirigido_a": g("dirigido_a", 255),
        "receptor": g("receptor", 255),
        "cargo": g("cargo", 255),
        "observacion": g("observacion"),
        "area": g("area", 255),
        "empresa": g("empresa", 255),
        "caducidad": None,
        "fecha_respuesta": g_date("fecha_respuesta"),
        "carta_respuesta": g("carta_respuesta", 255),
        "bandeja": g("bandeja", 80),
    }
    return base


def parse_cells_by_layout(layout: str, cells: tuple) -> dict | None:
    n_doc = _as_str(_val(cells, 3), 255)
    asunto = _as_str(_val(cells, 5))
    if not n_doc and not asunto:
        return None

    if n_doc and n_doc.upper() in ("TOTAL", "N°", "NRO", "N° DE DOCUMENTO", "DOCUMENTO"):
        return None

    base = {
        "n_orden": _as_int(_val(cells, 1)),
        "fecha": _as_date(_val(cells, 2)),
        "n_documento": n_doc or "S/N",
        "asunto": asunto,
        "especialidad": None,
        "estado": None,
        "referencias": None,
        "folios": None,
        "cd": None,
        "dirigido_a": None,
        "receptor": None,
        "cargo": None,
        "observacion": None,
        "area": None,
        "empresa": None,
        "caducidad": None,
        "fecha_respuesta": None,
        "carta_respuesta": None,
    }

    if layout == "emitida":
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "cd": _as_str(_val(cells, 10), 20),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
                "cargo": _as_str(_val(cells, 13), 255),
                "observacion": _as_str(_val(cells, 14)),
            }
        )
    elif layout == "recibida_sup":
        base.update(
            {
                "area": _as_str(_val(cells, 9), 255),
                "especialidad": _as_str(_val(cells, 10), 255),
                "estado": _as_str(_val(cells, 11), 120),
                "referencias": _as_str(_val(cells, 12)),
                "folios": _as_str(_val(cells, 13), 50),
                "cd": _as_str(_val(cells, 14), 20),
                "dirigido_a": _as_str(_val(cells, 15), 255),
                "receptor": _as_str(_val(cells, 16), 255),
                "cargo": _as_str(_val(cells, 17), 255),
                "fecha_respuesta": _as_date(_val(cells, 18)),
                "carta_respuesta": _as_str(_val(cells, 19), 255),
            }
        )
    elif layout == "recibida_simple":
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "observacion": _as_str(_val(cells, 10)),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
            }
        )
    elif layout == "recibida_otros":
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "cd": _as_str(_val(cells, 10), 20),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
                "cargo": _as_str(_val(cells, 13), 255),
                "observacion": _as_str(_val(cells, 14)),
            }
        )
    elif layout == "recibida_mpsc":
        base.update(
            {
                "especialidad": _as_str(_val(cells, 6), 255),
                "estado": _as_str(_val(cells, 7), 120),
                "referencias": _as_str(_val(cells, 8)),
                "folios": _as_str(_val(cells, 9), 50),
                "observacion": _as_str(_val(cells, 10)),
                "dirigido_a": _as_str(_val(cells, 11), 255),
                "receptor": _as_str(_val(cells, 12), 255),
            }
        )
        asunto_u = (base.get("asunto") or "").upper()
        if not base.get("estado") and any(
            k in asunto_u for k in ("HACE DE CONOCIMIENTO", "HACE LLEGAR COPIA", "AUTORIZACIÓN", "AUTORIZACION")
        ):
            base["estado"] = "PARA CONOCIMIENTO"
    else:
        return None

    base["estado_norm"] = infer_estado_from_row(base["estado"], base.get("asunto"), base.get("n_documento"), layout)
    if not base.get("estado") or base["estado"] == "SIN ESTADO":
        base["estado"] = base["estado_norm"]
    base["especialidad_norm"] = normalize_especialidad(base["especialidad"])
    return base


def _row_tuple(ban: str, sentido: str, row: dict) -> tuple:
    doc = row.get("n_documento") or "S/N"
    asunto = row.get("asunto")
    return (
        ban,
        sentido,
        row.get("n_orden"),
        row.get("fecha"),
        doc,
        infer_tipo_documento(doc, asunto),
        asunto,
        row.get("especialidad"),
        row.get("especialidad_norm"),
        row.get("estado"),
        row.get("estado_norm"),
        row.get("referencias"),
        row.get("folios"),
        row.get("cd"),
        row.get("dirigido_a"),
        row.get("receptor"),
        row.get("cargo"),
        row.get("observacion"),
        row.get("area"),
        row.get("empresa"),
        row.get("caducidad"),
        row.get("fecha_respuesta"),
        row.get("carta_respuesta"),
    )


def import_excel_to_db(conn, excel_path: Path | None = None, force: bool = False) -> dict:
    excel_path = Path(excel_path or os.environ.get("EXCEL_PATH") or DEFAULT_EXCEL)
    if not excel_path.exists():
        return {"ok": False, "error": f"Archivo Excel no encontrado: {excel_path}"}

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    sheetnames = list(wb.sheetnames)
    parsed_rows_by_bandeja: dict[str, list[tuple]] = {}
    total_parsed = 0
    sheets_processed = []

    try:
        matched_sheets = set()
        for cfg in SHEETS:
            actual_sheet_name = find_matching_sheet(sheetnames, cfg)
            if not actual_sheet_name or actual_sheet_name in matched_sheets:
                continue
            matched_sheets.add(actual_sheet_name)
            ws = wb[actual_sheet_name]
            header_row, col_map = detect_header_row_and_map(ws, cfg["header_row"])
            
            rows_for_sheet = []
            empty_streak = 0
            max_col = max(cfg["max_col"], max(col_map.values()) + 1 if col_map else 20)

            for cells in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
                if col_map:
                    parsed = parse_cells_with_map(cells or (), col_map)
                else:
                    parsed = parse_cells_by_layout(cfg["layout"], cells or ())

                if not parsed:
                    empty_streak += 1
                    if empty_streak >= EMPTY_STREAK_STOP:
                        break
                    continue
                empty_streak = 0
                rows_for_sheet.append(_row_tuple(cfg["bandeja"], cfg["sentido"], parsed))

            if rows_for_sheet:
                parsed_rows_by_bandeja[cfg["bandeja"]] = rows_for_sheet
                total_parsed += len(rows_for_sheet)
                sheets_processed.append(f"{actual_sheet_name} ({len(rows_for_sheet)} cartas)")

        # Si no se encontró ninguna de las 6 hojas estándar, o si hay hojas adicionales con cartas:
        if total_parsed == 0:
            for s_name in sheetnames:
                ws = wb[s_name]
                header_row, col_map = detect_header_row_and_map(ws, 1, max_scan=35)
                if not col_map and not any(k in s_name.lower() for k in ["carta", "doc", "res", "sup", "rl", "oficio", "informe"]):
                    continue
                
                rows_for_sheet = []
                empty_streak = 0
                max_col = max(col_map.values()) + 1 if col_map else 25

                s_low = s_name.lower()
                if "sup" in s_low or "recibida_sup" in s_low:
                    default_ban, default_sent = "recibida_sup", "recibida"
                elif "rl" in s_low or "legal" in s_low:
                    default_ban, default_sent = "rl", "emitida"
                elif "pronis" in s_low or "entidad" in s_low:
                    default_ban, default_sent = "recibida_pronis", "recibida"
                elif "mpsc" in s_low or "muni" in s_low:
                    default_ban, default_sent = "recibida_mpsc", "recibida"
                elif "otro" in s_low or "jrd" in s_low:
                    default_ban, default_sent = "recibida_otros", "recibida"
                else:
                    default_ban, default_sent = "residente", "emitida"

                for cells in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
                    if col_map:
                        parsed = parse_cells_with_map(cells or (), col_map)
                    else:
                        parsed = parse_cells_by_layout("emitida", cells or ())

                    if not parsed:
                        empty_streak += 1
                        if empty_streak >= EMPTY_STREAK_STOP:
                            break
                        continue
                    empty_streak = 0

                    raw_ban = str(parsed.get("bandeja") or "").lower()
                    rec = str(parsed.get("receptor") or "").upper()
                    if "sup" in raw_ban or "SUPERVIS" in rec:
                        b, s = "recibida_sup", "recibida"
                    elif "rl" in raw_ban or "LEGAL" in rec:
                        b, s = "rl", "emitida"
                    elif "pronis" in raw_ban or "PRONIS" in rec:
                        b, s = "recibida_pronis", "recibida"
                    elif "mpsc" in raw_ban or "MUNI" in rec:
                        b, s = "recibida_mpsc", "recibida"
                    elif "otro" in raw_ban or "JRD" in rec:
                        b, s = "recibida_otros", "recibida"
                    else:
                        b, s = default_ban, default_sent

                    rows_for_sheet.append(_row_tuple(b, s, parsed))

                if rows_for_sheet:
                    parsed_rows_by_bandeja[f"{s_name}_{default_ban}"] = rows_for_sheet
                    total_parsed += len(rows_for_sheet)
                    sheets_processed.append(f"{s_name} ({len(rows_for_sheet)} cartas)")

    finally:
        wb.close()

    # REGLA DE SEGURIDAD CRÍTICA
    if total_parsed == 0:
        return {
            "ok": False,
            "error": f"No se encontraron registros de cartas en el archivo Excel. Hojas disponibles: {sheetnames}. Asegúrate de que el archivo contenga las hojas de cartas o columnas como N° Documento, Fecha, Asunto.",
            "sheets": sheetnames,
        }

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) AS c FROM cartas")
        count = cur.fetchone()["c"]
        if count > 0 and not force:
            return {"ok": True, "skipped": True, "existing": count}

        # Borrar y reimportar de manera limpia y rápida
        cur.execute("SET FOREIGN_KEY_CHECKS=0")
        if force and count > 0:
            cur.execute("UPDATE cartas SET hilo_id=NULL")
            cur.execute("DELETE FROM cartas")
            cur.execute("DELETE FROM hilos")
            cur.execute("ALTER TABLE cartas AUTO_INCREMENT=1")
            cur.execute("ALTER TABLE hilos AUTO_INCREMENT=1")

        inserted = 0
        by_bandeja_summary = {}
        for ban_key, batch in parsed_rows_by_bandeja.items():
            if not batch:
                continue
            for i in range(0, len(batch), BATCH_SIZE):
                chunk = batch[i:i + BATCH_SIZE]
                cur.executemany(INSERT_SQL, chunk)
                inserted += len(chunk)
            by_bandeja_summary[ban_key] = {"inserted": len(batch)}
            print(f"[import] {ban_key}: {len(batch)} filas", flush=True)

        cur.execute("SET FOREIGN_KEY_CHECKS=1")
        try:
            fix_areas_responsables(conn, dry_run=False)
        except Exception as e:
            print(f"[import] Advertencia al normalizar áreas responsables: {e}", flush=True)

    conn.commit()
    return {
        "ok": True,
        "inserted": inserted,
        "by_bandeja": by_bandeja_summary,
        "sheets_processed": sheets_processed,
        "excel": str(excel_path),
    }
